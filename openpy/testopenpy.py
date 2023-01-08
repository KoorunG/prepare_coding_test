import collections

import openpyxl

filename = 'testxl.xlsx'
wb = openpyxl.load_workbook(filename)
sheet = wb['Sheet1']
num_dict = dict()
header = 'start'    # 초기화
data = 0            # 초기화
index = 10           # C열부터
while header:       # 헤더가 존재하는동안 반복
    num_dict[data] = header
    header = sheet.cell(row=4, column=index).value
    data = sheet.cell(row=5, column=index).value
    index += 1
print(num_dict[min(num_dict.keys())])